from django.contrib import admin
from .models import Attendance, Co_ordinator, Stage
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.utils import timezone
import pytz
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font, NamedStyle, Alignment, Border, Side


@admin.register(Attendance)
class AttendanceAdmin(admin.ModelAdmin):
    list_display = (
        "student_admn_no",
        "period",
        "time",
        "attendance",
        "co_ordinator",
        "stage",
        "program",
    )

    actions = ["download_excel"]

    def download_excel(self, request, queryset):
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="attendance_report.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active

        header_font = Font(bold=True)

        headers = [
            "Student Admission No",
            "Period",
            "Time",
            "Attendance",
            "Coordinator",
            "Stage",
            "Program",
        ]
        ist = pytz.timezone("Asia/Kolkata")
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f"{col_letter}1"].font = header_font
            worksheet[f"{col_letter}1"] = header

        # Define a custom style for the time column to adjust the width
        time_column_style = NamedStyle(
            name="time_column_style", number_format="yyyy-mm-ddThh:mm:ss"
        )

        # Set alignment for the header row
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            worksheet[f"{col_letter}1"].alignment = Alignment(
                horizontal="center", vertical="center"
            )

        for row_num, attendance in enumerate(queryset, 2):
            worksheet.cell(
                row=row_num, column=1, value=attendance.student_admn_no
            ).alignment = Alignment(horizontal="center", vertical="center")
            worksheet.cell(
                row=row_num, column=2, value=attendance.period
            ).alignment = Alignment(horizontal="center", vertical="center")

            time_ist = attendance.time.astimezone(ist).replace(tzinfo=None)
            worksheet.cell(
                row=row_num, column=3, value=time_ist
            ).style = time_column_style
            worksheet.cell(row=row_num, column=3).alignment = Alignment(
                horizontal="center", vertical="center"
            )

            worksheet.cell(
                row=row_num, column=4, value=attendance.attendance
            ).alignment = Alignment(horizontal="center", vertical="center")
            worksheet.cell(
                row=row_num, column=5, value=attendance.co_ordinator
            ).alignment = Alignment(horizontal="center", vertical="center")
            worksheet.cell(
                row=row_num, column=6, value=str(attendance.stage)
            ).alignment = Alignment(horizontal="center", vertical="center")
            worksheet.cell(
                row=row_num, column=7, value=str(attendance.program)
            ).alignment = Alignment(horizontal="center", vertical="center")

        # Set the width for the "Time" column
        worksheet.column_dimensions[
            get_column_letter(3)
        ].width = 20  # Adjust the width as needed

        workbook.save(response)

        return response


admin.site.register(Co_ordinator)
admin.site.register(Stage)
