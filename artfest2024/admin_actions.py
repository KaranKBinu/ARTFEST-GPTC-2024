# admin_actions.py
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.http import HttpResponse
from django.utils.encoding import smart_str
from .models import StudentUsers, Programs  # Import your models


def export_to_excel(modeladmin, request, queryset):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Create a mutable copy of the query parameters
    filters = request.GET.copy()

    # Remove empty or non-existent parameters
    filters = {key: value for key, value in filters.items() if value}

    # Check if there are any filters
    if filters:
        filter_headers = []
        program_code = ""
        program_name = ""

        # Convert program__id__exact to program (assuming program is the correct field name)
        filters = {
            key.replace("__id__exact", ""): value for key, value in filters.items()
        }
        for key, value in filters.items():
            if key == "program":
                try:
                    program_code = Programs.objects.get(id=value).program_code
                    program_name = Programs.objects.get(id=value).program_name
                    filter_headers.append(f"{key}: {program_name}")
                except Programs.DoesNotExist:
                    filter_headers.append(f"{key}: {value}")
            else:
                filter_headers.append(f"{key}: {value}")

        # Create a unique filename based on the selected filters
        filename_filters = "_".join(f"{key}_{value}" for key, value in filters.items())
        filename = (
            f"{program_code}_{filename_filters}.xlsx"
            if program_code
            else f"students_data_{filename_filters}.xlsx"
        )

        response["Content-Disposition"] = f"attachment; filename={smart_str(filename)}"

        workbook = Workbook()
        worksheet = workbook.active

        # Append the filter headers to the first row
        worksheet.append(filter_headers)

        # Add standard headers to the worksheet
        standard_headers = [
            "Admission Number",
            "Name",
            "Email",
            "Phone",
            "Password",
            "House Name",
            "Gender",
            "Programs",
        ]

        # Append the standard headers after the first row
        worksheet.append(standard_headers)

        # Add data to the worksheet
        for student in queryset:
            row_data = [
                student.student_admn_no,
                student.student_name,
                student.student_email,
                student.student_phone,
                student.student_password,
                student.house_name,
                student.Gender,
                ", ".join(str(program) for program in student.program.all()),
            ]
            worksheet.append(row_data)

        # Adjust column widths and apply alignment
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            for cell in column:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

        # Set the header row to span all columns
        header_row_range = (
            f"A1:{chr(ord('A') + len(filter_headers) + len(standard_headers) - 1)}1"
        )
        worksheet.merge_cells(header_row_range)

        workbook.save(response)
        return response

    else:
        # If no filters are applied, use the existing code with a custom heading
        filter_headers = []
        program_code = ""
        program_name = ""

        filename = "students_data_all.xlsx"

        response["Content-Disposition"] = f"attachment; filename={smart_str(filename)}"

        workbook = Workbook()
        worksheet = workbook.active

        # Add a custom heading for all students
        all_students_heading = ["All Students"]
        worksheet.append(all_students_heading)

        # Add standard headers to the worksheet
        standard_headers = [
            "Admission Number",
            "Name",
            "Email",
            "Phone",
            "Password",
            "House Name",
            "Gender",
            "Programs",
        ]

        # Append the standard headers to the first row
        worksheet.append(standard_headers)

        # Add data to the worksheet
        for student in queryset:
            row_data = [
                student.student_admn_no,
                student.student_name,
                student.student_email,
                student.student_phone,
                student.student_password,
                student.house_name,
                student.Gender,
                ", ".join(str(program) for program in student.program.all()),
            ]
            worksheet.append(row_data)

        # Adjust column widths and apply alignment
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            for cell in column:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

        # Set the header row to span all columns
        header_row_range = f"A1:{chr(ord('A') + len(all_students_heading) + len(standard_headers) - 1)}1"
        worksheet.merge_cells(header_row_range)

        workbook.save(response)
        return response


export_to_excel.short_description = "Export selected students to Excel"
