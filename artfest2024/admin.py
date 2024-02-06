from django.contrib import admin
from .models import GroupProgram, Programs, StudentDetails, StudentUsers, ContactMessage
from django.contrib import admin
from .models import StudentUsers, Notification
from .admin_actions import export_to_excel
from openpyxl.utils import get_column_letter  # Add this import statement

from django.http import HttpResponse
from openpyxl import Workbook
import csv

# admin.py


from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


@admin.register(StudentDetails)
class StudentDetailsAdmin(admin.ModelAdmin):
    list_display = ["student_name", "admn_no", "branch", "sem"]
    list_filter = ["house", "branch", "sem"]
    search_fields = ["student_name", "admn_no"]

    actions = ["download_excel"]

    def download_excel(self, request, queryset):
        # Disable pagination
        self.list_per_page = self.get_queryset(request).count()

        # Create a workbook and add a worksheet.
        wb = Workbook()
        ws = wb.active

        # Write headers to the worksheet and adjust formatting
        headers = ["Student Name", "Admission No", "Branch", "Semester", "House Name"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            # Set alignment to center
            ws[f"{col_letter}1"].alignment = Alignment(horizontal="center")

        # Write data to the worksheet and format cells
        for row_num, student in enumerate(queryset, 2):
            ws[f"A{row_num}"] = student.student_name
            ws[f"B{row_num}"] = student.admn_no
            ws[f"C{row_num}"] = student.branch
            ws[f"D{row_num}"] = student.sem
            ws[f"E{row_num}"] = student.house

            # Set alignment to center for each cell
            for col_num in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}{row_num}"].alignment = Alignment(horizontal="center")

        # Reset pagination to default
        self.list_per_page = 1000

        # Create response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=student_details.xlsx"
        wb.save(response)

        return response

    download_excel.short_description = "Download Excel"


class StudentUsersAdmin(admin.ModelAdmin):
    list_display = (
        "student_admn_no",
        "student_name",
        "student_email",
        "student_phone",
        "house_name",
        "Gender",
    )
    list_filter = (
        "house_name",
        "Gender",
        "program",
    )  # Add these lines to enable filtering
    actions = [export_to_excel]


admin.site.register(StudentUsers, StudentUsersAdmin)
admin.site.register(GroupProgram)

admin.site.register(Programs)


class ContactMessageAdmin(admin.ModelAdmin):
    list_display = ("name", "email", "subject")


admin.site.register(ContactMessage, ContactMessageAdmin)


admin.site.register(Notification)
